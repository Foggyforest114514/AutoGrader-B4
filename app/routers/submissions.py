from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional
from datetime import datetime
import uuid
import io
from app.database import get_db
from app.models.models import User, Submission, Assignment, ClassStudent, Student
from app.schemas import SubmissionCreate, SubmissionResultUpdate, ResponseModel
from app.auth import get_current_user

router = APIRouter(prefix="/api/v1/submissions", tags=["提交管理"])

@router.post("", response_model=ResponseModel)
async def create_submission(
    submission_data: SubmissionCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role != "student":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有学生可以提交作业"
        )
    
    assignment = db.query(Assignment).filter(
        Assignment.assignment_id == submission_data.assignment_id
    ).first()
    
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="作业不存在"
        )
    
    if not assignment.is_published:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="作业未发布"
        )
    
    student_in_class = db.query(ClassStudent).filter(
        ClassStudent.class_id == assignment.class_id,
        ClassStudent.student_user_id == current_user.user_id
    ).first()
    
    if not student_in_class:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您不在该作业所在的班级中"
        )
    
    if datetime.utcnow() > assignment.due_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="作业已截止"
        )
    
    submission_id = str(uuid.uuid4())
    
    new_submission = Submission(
        submission_id=submission_id,
        student_user_id=current_user.user_id,
        question_id=submission_data.question_id,
        assignment_id=submission_data.assignment_id,
        code=submission_data.code,
        language=submission_data.language,
        status="PENDING"
    )
    
    db.add(new_submission)
    db.commit()
    db.refresh(new_submission)
    
    return ResponseModel(
        code=200,
        msg="提交成功",
        data={"submission_id": submission_id}
    )

@router.get("/my", response_model=ResponseModel)
async def get_my_submissions(
    assignment_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(Submission).options(
        joinedload(Submission.assignment)
    ).filter(
        Submission.student_user_id == current_user.user_id
    )

    if assignment_id:
        query = query.filter(Submission.assignment_id == assignment_id)

    submissions = query.order_by(Submission.submitted_at.desc()).all()

    submissions_data = []
    for submission in submissions:
        submission_dict = {
            "submission_id": submission.submission_id,
            "assignment_id": submission.assignment_id,
            "assignment_title": submission.assignment.title,
            "question_id": submission.question_id,
            "submitted_at": submission.submitted_at.isoformat(),
            "status": submission.status,
            "overall_score": submission.overall_score,
            "passed_count": submission.passed_count,
            "total_count": submission.total_count
        }
        submissions_data.append(submission_dict)

    return ResponseModel(
        code=200,
        msg="成功",
        data=submissions_data
    )

@router.get("/{submission_id}", response_model=ResponseModel)
async def get_submission_detail(
    submission_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    submission = db.query(Submission).options(
        joinedload(Submission.assignment),
        joinedload(Submission.student)
    ).filter(
        Submission.submission_id == submission_id
    ).first()
    
    if not submission:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="提交记录不存在"
        )
    
    if current_user.role == "student" and submission.student_user_id != current_user.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限查看此提交记录"
        )
    
    submission_dict = {
        "submission_id": submission.submission_id,
        "student_user_id": submission.student_user_id,
        "student_name": submission.student.real_name,
        "question_id": submission.question_id,
        "assignment_id": submission.assignment_id,
        "assignment_title": submission.assignment.title,
        "code": submission.code,
        "language": submission.language,
        "submitted_at": submission.submitted_at.isoformat(),
        "status": submission.status,
        "overall_score": submission.overall_score,
        "passed_count": submission.passed_count,
        "total_count": submission.total_count,
        "overall_comment": submission.overall_comment,
        "static_issues": submission.static_issues,
        "case_results": submission.case_results,
        "teacher_score_override": submission.teacher_score_override,
        "override_reason": submission.override_reason
    }
    
    return ResponseModel(
        code=200,
        msg="成功",
        data=submission_dict
    )

@router.patch("/{submission_id}/result", response_model=ResponseModel)
async def update_submission_result(
    submission_id: str,
    result_data: SubmissionResultUpdate,
    db: Session = Depends(get_db)
):
    submission = db.query(Submission).filter(
        Submission.submission_id == submission_id
    ).first()

    if not submission:
        if result_data.studentUserId and result_data.assignmentId and result_data.code:
            submission = Submission(
                submission_id=submission_id,
                student_user_id=result_data.studentUserId,
                assignment_id=result_data.assignmentId,
                question_id=result_data.questionId or "",
                code=result_data.code,
                language=result_data.language or "python",
                status="PENDING",
            )
            db.add(submission)
            db.flush()
        else:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="提交记录不存在，且缺少创建所需字段"
            )
    
    submission.status = result_data.status
    
    if result_data.overallScore is not None:
        submission.overall_score = result_data.overallScore
    
    if result_data.passedCount is not None:
        submission.passed_count = result_data.passedCount
    
    if result_data.totalCount is not None:
        submission.total_count = result_data.totalCount
    
    if result_data.overallComment is not None:
        submission.overall_comment = result_data.overallComment
    
    if result_data.staticIssues is not None:
        submission.static_issues = result_data.staticIssues
    
    if result_data.caseResults is not None:
        submission.case_results = result_data.caseResults
    
    db.commit()
    db.refresh(submission)
    
    return ResponseModel(
        code=200,
        msg="评测结果更新成功",
        data={"submission_id": submission_id}
    )

@router.get("/assignment/{assignment_id}/all", response_model=ResponseModel)
async def get_assignment_all_submissions(
    assignment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    assignment = db.query(Assignment).filter(
        Assignment.assignment_id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="作业不存在"
        )

    if current_user.role == "teacher" and assignment.teacher_id != current_user.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限查看此作业的提交"
        )

    submissions = db.query(Submission).options(
        joinedload(Submission.student)
    ).filter(
        Submission.assignment_id == assignment_id
    ).order_by(Submission.submitted_at.desc()).all()

    # 批量获取学号映射，避免 N+1
    student_user_ids = list(set(s.student_user_id for s in submissions))
    student_id_map = {}
    if student_user_ids:
        student_records = db.query(Student).filter(Student.user_id.in_(student_user_ids)).all()
        student_id_map = {s.user_id: s.student_id for s in student_records}

    submissions_data = []
    for submission in submissions:
        submission_dict = {
            "submission_id": submission.submission_id,
            "student_user_id": submission.student_user_id,
            "student_name": submission.student.real_name,
            "student_id": student_id_map.get(submission.student_user_id, ''),
            "question_id": submission.question_id,
            "submitted_at": submission.submitted_at.isoformat(),
            "status": submission.status,
            "overall_score": submission.overall_score,
            "passed_count": submission.passed_count,
            "total_count": submission.total_count
        }
        submissions_data.append(submission_dict)

    return ResponseModel(
        code=200,
        msg="成功",
        data={
            "assignment_title": assignment.title,
            "total_submissions": len(submissions_data),
            "submissions": submissions_data
        }
    )

@router.patch("/{submission_id}/override", response_model=ResponseModel)
async def override_submission_score(
    submission_id: str,
    override_score: float,
    override_reason: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role not in ["teacher", "admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有教师或管理员可以覆盖分数"
        )

    submission = db.query(Submission).filter(
        Submission.submission_id == submission_id
    ).first()

    if not submission:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="提交记录不存在"
        )

    assignment = db.query(Assignment).filter(
        Assignment.assignment_id == submission.assignment_id
    ).first()

    if current_user.role == "teacher" and assignment.teacher_id != current_user.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只能覆盖自己布置的作业的分数"
        )

    submission.teacher_score_override = override_score
    submission.override_reason = override_reason

    db.commit()
    db.refresh(submission)

    return ResponseModel(
        code=200,
        msg="分数覆盖成功",
        data={
            "submission_id": submission_id,
            "override_score": override_score
        }
    )

@router.get("/statistics/assignment/{assignment_id}", response_model=ResponseModel)
async def get_assignment_submission_statistics(
    assignment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    assignment = db.query(Assignment).filter(
        Assignment.assignment_id == assignment_id
    ).first()

    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="作业不存在"
        )

    if current_user.role == "teacher" and assignment.teacher_id != current_user.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权限查看此作业的统计"
        )

    submissions = db.query(Submission).filter(
        Submission.assignment_id == assignment_id,
        Submission.status == "COMPLETED"
    ).all()

    total_submissions = len(submissions)
    total_students = db.query(ClassStudent).filter(
        ClassStudent.class_id == assignment.class_id
    ).count()

    # 去重：按提交学生学号计算，避免同一学生多次提交导致提交率超过100%
    submitted_student_ids = set(s.student_user_id for s in submissions)
    submitted_student_count = len(submitted_student_ids)

    scores = [s.overall_score for s in submissions if s.overall_score is not None]
    average_score = sum(scores) / len(scores) if scores else 0
    max_score = max(scores) if scores else 0
    min_score = min(scores) if scores else 0

    passed_count = len([s for s in submissions if s.overall_score and s.overall_score >= 60])
    pass_rate = (passed_count / total_submissions * 100) if total_submissions > 0 else 0

    return ResponseModel(
        code=200,
        msg="成功",
        data={
            "assignment_id": assignment_id,
            "assignment_title": assignment.title,
            "total_students": total_students,
            "total_submissions": total_submissions,
            "submission_rate": (submitted_student_count / total_students * 100) if total_students > 0 else 0,
            "average_score": round(average_score, 2),
            "max_score": max_score,
            "min_score": min_score,
            "pass_rate": round(pass_rate, 2),
            "score_distribution": {
                "excellent": len([s for s in scores if s >= 90]),
                "good": len([s for s in scores if 80 <= s < 90]),
                "medium": len([s for s in scores if 70 <= s < 80]),
                "pass": len([s for s in scores if 60 <= s < 70]),
                "fail": len([s for s in scores if s < 60])
            }
        }
    )


@router.get("/export/assignment/{assignment_id}", response_class=StreamingResponse)
async def export_submissions_excel(
    assignment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出指定作业的所有提交成绩为 Excel 文件"""
    if current_user.role not in ["teacher", "admin"]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅教师和管理员可导出")

    assignment = db.query(Assignment).filter(Assignment.assignment_id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="作业不存在")

    submissions = db.query(Submission).filter(
        Submission.assignment_id == assignment_id
    ).order_by(Submission.submitted_at.desc()).all()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="openpyxl 未安装")

    wb = Workbook()
    ws = wb.active
    ws.title = "提交成绩"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["提交ID", "学生名", "学号", "题目ID", "语言", "状态", "总分", "通过/总数", "评语", "提交时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    for row_idx, s in enumerate(submissions, 2):
        student = db.query(User).filter(User.user_id == s.student_user_id).first()
        student_info = db.query(Student).filter(Student.user_id == s.student_user_id).first()
        student_name = student.real_name if student else str(s.student_user_id)
        student_id = student_info.student_id if student_info else str(s.student_user_id)

        row_data = [
            s.submission_id,
            student_name,
            student_id,
            s.question_id or "",
            s.language or "",
            s.status or "",
            s.overall_score or 0,
            f"{s.passed_count or 0}/{s.total_count or 0}",
            s.overall_comment or "",
            s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col in (1, 6, 7):
                cell.alignment = Alignment(horizontal="center")

    # 列宽
    widths = [38, 12, 14, 12, 10, 12, 8, 10, 40, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"assignment_{assignment_id}_{assignment.title}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
